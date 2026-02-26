#!/usr/bin/env python3
"""
Export each sheet of an Excel workbook to a separate CSV in the same folder.
Usage: python scripts/export_xlsx_sheets_to_csv.py [path_to.xlsx]
Default: data_reference/Air_Monitoring_Relationships-2.xlsx
"""
import csv
import os
import sys

def main():
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
    else:
        xlsx_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "data_reference", "Air_Monitoring_Relationships-2.xlsx"
        )
    if not os.path.isfile(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("Need openpyxl. Run: pip install openpyxl")
        sys.exit(1)

    out_dir = os.path.dirname(xlsx_path)
    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        safe = "".join(c if c.isalnum() or c in " -_" else "_" for c in sheet_name).strip() or "sheet"
        out_path = os.path.join(out_dir, f"{base}_{safe}.csv")
        ws = wb[sheet_name]
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else str(v) for v in row])
        print(out_path)
    wb.close()
    print("Done.")

if __name__ == "__main__":
    main()
